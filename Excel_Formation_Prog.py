import openpyxl
from openpyxl import Workbook
from googletrans import Translator
# from openpyxl.styles import PatternFill

# a = input("Do you already have excel file completed|| yes(1) , no(2) :" )
# director = int(a)
# if director == 1 :
# print("\nENGLISH ADDRESS INFO:\n")
# a = input(" Enter the column of English address : ")
eng_adr_col = int(12)
# a = input("Enter the row from the English address starts :")
str_row_eng_adr = int (2)

# a = input("Enter the Englsih Tithi column :")
tithi_col = int(11)
# a = input("Enter the row from the English Tithi starts :")
str_row_tithi = int(2)
# print("______________________________INSTRUCTIONS______________________________")
# print("\n\n\n 1) Kindly take 1st sheet as an Datewise display sheet.\n\n\n")
# print("\n\n 2) Give datewise sheet name as 'S1'. (i.e.: Capital S1)\n\n")
# print("\n\n 3) Don't Open the destination file through out this programme.\n\n")
# print("\n\n 4) Starts our sheet actual data from 2nd row.\n\n\n")
print("_______________________________EXCEL SHEET PROGRAMME STARTS__________________________________")
locn_excel  = input("\n\nEnter the exact location with file formate :")
dest_loc_excel = input("Enter the destination location for new excel workbook is going to creat : ")
wb = openpyxl.load_workbook(locn_excel)

date_col = 10
eng_name_col = 8
guj_name_col = 9

sheet1 = wb['S1']
i = 1
while sheet1.cell(row= i ,column= date_col).value != None:
    i = i + 1

max_rox = i
max_col = sheet1.max_column

def tithi_eng(tithi , row):
    full_eg_tt_dic =  { "B.D." : "Birthday" , "P.T." : "Punyatithi" , "Brh. Bhoj."  : "Brahm Bhojan" , "E.A." : "Earning Anniversary" , "M.A." : "Marriage Anniversary" , "Brh.Bhoj." : "Brahm Bhojan"}
    
    if  tithi == "B.D.": 
        sheet1.cell(row= row , column=(max_col + 4)).value  = full_eg_tt_dic[tithi]
    
    elif  tithi ==  "P.T.": 
        sheet1.cell(row= row , column=(max_col + 4)).value  = full_eg_tt_dic[tithi]
    
    elif  tithi == "Brh. Bhoj.": 
        sheet1.cell(row= row , column=(max_col + 4)).value  = full_eg_tt_dic[tithi]
    
    elif  tithi == "E.A." : 
        sheet1.cell(row= row , column=(max_col + 4)).value  = full_eg_tt_dic[tithi]

    elif tithi == "M.A.":
        sheet1.cell(row= row , column=(max_col + 4)).value  = full_eg_tt_dic[tithi]

    else : 
        sheet1.cell(row= row , column=(max_col + 4)).value  = None


def tithi_guj(tithi,temprow):
    if tithi == "B.D.":
        sheet1.cell(row= temprow, column= (max_col + 1)).value = "જન્મદિવસ"
        
    elif tithi == "P.T.":
        sheet1.cell(row= temprow, column= (max_col + 1)).value = "પુણ્યતિથિ"
        

    elif tithi == "Brh. Bhoj.":
        sheet1.cell(row= temprow, column= (max_col + 1)).value  = "બ્રહ્મભોજન​"

    elif tithi == "E.A.":
        sheet1.cell(row= temprow, column= (max_col + 1)).value  = "કમાણીની વર્ષગાંઠ​"
    
    elif tithi == "M.A.":
        sheet1.cell(row= temprow, column= (max_col + 1)).value  = "લગ્નદિન​​"

    else:
        sheet1.cell(row= temprow, column= (max_col + 1)).value  = None


def adr_guj(adr,row):
    if adr == "Amdavad":
        sheet1.cell(row= row, column= (max_col + 2)).value = "અમદાવાદ​"

    elif adr == "Andheri":
        sheet1.cell(row= row, column= (max_col + 2)).value = "અંધેરી​"

    elif adr == "Bardoli":
        sheet1.cell(row= row, column= (max_col + 2)).value = "બારડોલી"
    
    elif adr == "Mahesana":
        sheet1.cell(row= row, column= (max_col + 2)).value = "મહેસાણા"
    
    elif adr == "Thane":
        sheet1.cell(row= row, column= (max_col + 2)).value = "ઠાણે"

    elif adr == "Chembur":
        sheet1.cell(row= row, column= (max_col + 2)).value = "ચેમ્બૂર​"

    elif adr == "Ankleshwar" or adr == "Ankleshvar":
        sheet1.cell(row= row, column= (max_col + 2)).value = "અંક્લેશ્વર​​​​"

    elif adr == "Bhavnagar":
        sheet1.cell(row= row, column= (max_col + 2)).value = "ભાવનગર​"

    elif adr == "Bharuch":
        sheet1.cell(row= row, column= (max_col + 2)).value = "ભરુચ​​"

    elif adr == "Kandivali":
        sheet1.cell(row= row, column= (max_col + 2)).value = "કાંદિવલી​"

    elif adr == "Navi Mumbai":
        sheet1.cell(row= row, column= (max_col + 2)).value = "ન​વી મુંબઈ​"

    elif adr == "Mumbai":
        sheet1.cell(row= row, column= (max_col + 2)).value = "મુંબઈ​"

    elif adr == "Pune":
        sheet1.cell(row= row, column= (max_col + 2)).value = "પુણે"

    elif adr == "Vashi":
        sheet1.cell(row= row, column= (max_col + 2)).value = "વાશી"

    elif adr == "Vadodara":
        sheet1.cell(row= row, column= (max_col + 2)).value = "વડોદરા"

    elif adr == "Borivali":
        sheet1.cell(row= row, column= (max_col + 2)).value = "બોરિવલી"

    elif adr == "Bandra":
        sheet1.cell(row= row, column= (max_col + 2)).value = "બાન્દ્રા"

    elif adr == "Chennai":
        sheet1.cell(row= row, column= (max_col + 2)).value = "ચેન્નઈ"

    elif adr == "Powai":
        sheet1.cell(row= row, column= (max_col + 2)).value = "પ​વઈ"

    elif adr == "Surat":
        sheet1.cell(row= row, column= (max_col + 2)).value = "સુરત​"

    # if adr == "Ankleshwar" or adr == "Ankleshvar":
    #     sheet1.cell(row= row, column= (max_col + 2)).value = "અંક્લેશ્વર​​​​"

    elif adr == "Mahim":
        sheet1.cell(row= row, column= (max_col + 2)).value = "માહીમ​"

    elif adr == "Goregaon":
        sheet1.cell(row= row, column= (max_col + 2)).value = "ગોરેગાંવ​"

    elif adr == "Vile Parle":
        sheet1.cell(row= row, column= (max_col + 2)).value = "વિલે પાર્લે"

    elif adr == "Santacruz":
        sheet1.cell(row= row, column= (max_col + 2)).value = "સાંતાક્રુઝ​​"

    elif adr == "Matunga":
        sheet1.cell(row= row, column= (max_col + 2)).value = "માટુંગા"

    elif adr == "Ghatkoper" or adr == "Ghatkopar":
        sheet1.cell(row= row, column= (max_col + 2)).value = "ઘાટકોપર​"

    elif adr == "Valsad":
        sheet1.cell(row= row, column= (max_col + 2)).value = "વલસાડ"

    elif adr == "Pardi":
        sheet1.cell(row= row, column= (max_col + 2)).value = "પારડી​"

    elif adr == "Khar":
        sheet1.cell(row= row, column= (max_col + 2)).value = "ખાર​"

    elif adr == "Mulund" or adr == "Muland":
        sheet1.cell(row= row, column= (max_col + 2)).value = "મુલુંડ​"

    elif adr == "Parel":
        sheet1.cell(row= row, column= (max_col + 2)).value = "પરેલ​​"
        
    elif adr == "Malad":
        sheet1.cell(row= row, column= (max_col + 2)).value = "માલાડ​"

    elif adr == "Juhu":
        sheet1.cell(row= row, column= (max_col + 2)).value = "જુહૂ​"

    elif adr == "Dadar":
        sheet1.cell(row= row, column= (max_col + 2)).value = "દાદર​"

    elif adr == "Kapadvanj":
        sheet1.cell(row= row, column= (max_col + 2)).value = "કાપડ​વંજ​"

    elif adr == "Walkeshwar":
        sheet1.cell(row= row, column= (max_col + 2)).value = "વાલકેશ્વર​"

    elif adr == "U.S.A.":
        sheet1.cell(row= row, column= (max_col + 2)).value = "યુ.એસ​.એ."

    elif adr == "Bhayandar" or adr == "Bhayender":
        sheet1.cell(row= row, column= (max_col + 2)).value = "ભાયંદર​"

    elif adr == "Killa Pardi":
        sheet1.cell(row= row, column= (max_col + 2)).value = "કિલ્લા પારડી"

    elif adr == "Bhayli":
        sheet1.cell(row= row, column= (max_col + 2)).value = "ભૈલી"

    elif adr == "Jogeshwari":
        sheet1.cell(row= row, column= (max_col + 2)).value = "જોગેશ્વરી"

    elif adr == "Kosamba":
        sheet1.cell(row= row, column= (max_col + 2)).value = "કોસંબા"

    elif adr == "London":
        sheet1.cell(row= row, column= (max_col + 2)).value = "લંડન​"
    
    elif adr == "Coimbatore":
        sheet1.cell(row= row, column= (max_col + 2)).value = "કોયંબત્તૂર​​"

    elif adr == "Kings Circle":
        sheet1.cell(row= row, column= (max_col + 2)).value = "કિંગ્સ સર્કલ​"

    elif adr == "Nasik":
        sheet1.cell(row= row, column= (max_col + 2)).value = "નાસિક​"

    elif adr == "Sion":
        sheet1.cell(row= row, column= (max_col + 2)).value = "સાયન​"

    elif adr == "Vapi":
        sheet1.cell(row= row, column= (max_col + 2)).value = "વાપી"

    elif adr == "Valia":
        sheet1.cell(row= row, column= (max_col + 2)).value = "વાલિયા"

    elif adr == "Udvada":
        sheet1.cell(row= row, column= (max_col + 2)).value = "ઉદ​વાદા"

    elif adr == "Nipani":
        sheet1.cell(row= row, column= (max_col + 2)).value = "નિપાની"

    elif adr == "Salem":
        sheet1.cell(row= row, column= (max_col + 2)).value = "સેલમ​"

    elif adr == "Vidhyavihar":
        sheet1.cell(row= row, column= (max_col + 2)).value = "વિદ્યાવિહાર​"

    else:
        sheet1.cell(row= row, column= (max_col + 2)).value  = None


def date_guj(date1):

    
    day = { "01" : "૧ ","02" : "૨ ", "03" : "૩ ", "04" : "૪ " , "05" : "૫ ", "06" : "૬ ", "07" : "૭ " , "08" : "૮ ", "09" : "૯ ", "10" : "૧૦","11" : "૧૧ ","12" : "૧૨ ", "13" : "૧૩ ", "14" : "૧૪ " , "15" : "૧૫ ", "16" : "૧૬ ", "17" : "૧૭ " , "18" : "૧૮ ", "19" : "૧૯ ", "20" : "૨૦ ", "21" : "૨૧ ","22" : "૨૨ ", "23" : "૨૩ ", "24" : "૨૪ " , "25" :"૨૫ ", "26" : "૨૬ ", "27" : "૨૭ " , "28" : "૨૮ ", "29" : "૨૯ ", "30" : "૩૦ ","31" : "૩૧ "}
    
    date_dig = day[date1]

    return date_dig
    


def fn_guj_month(given):
    months = {"01" : "જાન્યુઆરી", "02" : "ફેબ્રુઆરી" , "03" : "માર્ચ​" ,"04" : "એપ્રિલ​", "05" : "મે" , "06" : "જૂન" , "07" : "જુલાઈ" , "08" : "ઓગસ્ટ" , "09" : "સપ્ટેમ્બર" , "10" : "ઓક્ટોબર" , "11" : "નવેમ્બર", "12" : "ડિસેમ્બર"}
    guj_month_name = months[given]

    return guj_month_name

    
r = str_row_tithi
while r < max_rox + 1:
    print(r)
    run_tithi = sheet1.cell(row= r ,column = tithi_col).value
    tithi_guj(run_tithi,r)
    r = r + 1

print("\n\n -> All Applicable Tithies have been printed in gujrati.")

r = str_row_tithi
while r < max_rox + 1:
    print(r)
    run_tithi = sheet1.cell(row= r ,column = tithi_col).value
    tithi_eng(run_tithi ,r )
    r = r + 1

r = str_row_eng_adr
while r < max_rox +1:
    run_adr = sheet1.cell(row= r, column= eng_adr_col).value
    adr_guj(run_adr,r)
    r = r + 1

print("\n -> All Applicable address have benn converted in english ")

# date_col = int(input("Enter the English date column : "))

# str_row_date = int(input("Enter the row from date starts : "))
str_row_date = int(2)
# print("\nPlease Input only dates upto which Eng dates were there :) \n")
# end_row_date = int(input("Enter the row upto GUJ date prints : "))


r = 2
print(max_rox)
while r < max_rox:
    run_date = str(sheet1.cell(row= r, column= date_col).value)
    if run_date == "None":
        break
    temlist = list(run_date)
    day = temlist [8] + temlist[9]
    curr_month = temlist[5] + temlist[6]
    guj_month_name = fn_guj_month(curr_month)
    guj_digit = date_guj(day)
    actual_name = guj_digit + guj_month_name
    sheet1.cell(row = r ,column= (max_col + 3) ).value = actual_name
    r= r+1

print("\nAll compitible Gujrati dates have been Printed...")
 
r = 2
translator = Translator()
while r < (max_rox +1 ):
    print(r)
    run_name = str(sheet1.cell(row= r , column=eng_name_col ).value)
    if run_name.find('Chi') == 0:
        a = run_name.split()
        if len(a) == 4:
            act_name = a[1] +" "+ a[2] + " " + a[3]
            trns_act_name = translator.translate(act_name, scr = 'en' , dest= 'gu').text
            f_guj_name = "ચિ." + " " + trns_act_name
            print(f_guj_name)
            sheet1.cell(row=r, column= guj_name_col).value = f_guj_name
        elif len(a) == 3:
            act_name = a[1] +" "+ a[2] 
            trns_act_name = translator.translate(act_name, scr = 'en' , dest= 'gu').text
            f_guj_name = "ચિ." + " " + trns_act_name
            print(f_guj_name)
            sheet1.cell(row=r, column= guj_name_col).value = f_guj_name

    else:
        translated = translator.translate(run_name  , src='en' , dest= 'gu').text
        sheet1.cell(row=r, column= guj_name_col).value = translated
        print(translated)
    r = r + 1
    print(r)
print("\n All The appropriate names as per this have been Printed in sheet...")

wb.save(dest_loc_excel)
print("\n\n\n___________________~Your new excel file have been saved.~____________________   :) \n\n\n")

# print("\n\n\n__________________Presantation making process is going to start.")
# import collections.abc
# from pptx import Presentation

# print("_____________________INSTRUCTION_____________________")
# print("\n 1) Make sure that your Model.pptx has a copy at this location.(C:\Model.pptx)")
# print("\n 2) All the English Names should be in the formate like(Shri. Ombhai R. Patel):")
# print("______________________________________________________________________")
# loc_ppt = "C:\\Model.F.pptx"
# desti_loc_ppt = input("\n\n\nEnter the destination location of new ppt is going to creat : ")

# last_row_date = max_rox 
# # print("_________________COLUMN INFORMATION______________________")
# ppt_guj_name_col = int(9)
# ppt_eng_name_col = int(8)
# ppt_tithi_guj_col = int(13)
# ppt_adr_eng_col = int(12)
# ppt_adr_guj_col = int(14)
# # ppt_tithi_eng_col = int(input("Enter the English Tithi's Column : "))
# ppt_date_guj_col = int(15)
# ppt_tithi_eng_col = int(16)

# r = int(2)

# #_____PPT Making starts

# prs = Presentation(loc_ppt)


# # guj_tithi_col = int(max_col) + 1
# # guj_adr_col = int(max_col) + 2
# # # guj_date_col = int(max_col) + 3

# sld_no = int(0)
# while sld_no <= max_rox  :
#     run_slide = prs.slides[sld_no]
#     shapes = run_slide.shapes
#     r = sld_no + 2
#     print(sld_no)

#     full_date = sheet1.cell(row = r, column= ppt_date_guj_col ).value
#     if type(full_date) == None:
#         pass

#     else :
#         tf_guj_date = shapes[4].text_frame.paragraphs[0]
#         fd_sp = full_date.split()
#         d_runs0 = fd_sp[0]
#         d_runs2 = fd_sp[1]
#         tf_guj_date.runs[0].text = d_runs0
#         tf_guj_date.runs[2].text = d_runs2

    

#     a =     sheet1.cell(row = r,column= ppt_guj_name_col).value
#     if type(a) == None:
#         pass

#     else:
#         tf_guj_name = shapes[5].text_frame.paragraphs[0]
#         tf_guj_name.runs[0].text = sheet1.cell(row = r,column= ppt_guj_name_col).value

#     #ENG_NAME::::
#     tf_eng_name = shapes[6].text_frame.paragraphs[0]
#     full_name = sheet1.cell(row = r , column= ppt_eng_name_col).value

#     if type(full_name) == None:
#         pass
#     else:
#         ful_n_split = full_name.split()
        
#         l = len(ful_n_split)
#         if l == 4:
#             runs0  = ful_n_split[0]
#             runs1  = " " +ful_n_split [1]
#             runs2  = " " + ful_n_split[2] + " " + ful_n_split[3]


#             tf_eng_name.runs[0].text = runs0
#             tf_eng_name.runs[1].text = runs1
#             tf_eng_name.runs[2].text = runs2
        
#         else:
#             # fill_cell = PatternFill(patternType='solid',fgColor='ff4d4d')
#             # sheet1.cell(row= r, column=ppt_eng_name_col ).fill = fill_cell
#             pass

#     a = sheet1.cell(row= r , column= ppt_tithi_guj_col).value
#     if type(a) == None:
#         pass
#     else:
#         tf_tithi_guj = shapes[7].text_frame.paragraphs[0]
#         tf_tithi_guj.runs[0].text = sheet1.cell(row= r , column= ppt_tithi_guj_col).value

#     a = sheet1.cell(row= r , column= ppt_adr_eng_col).value
#     if type(a) == None:
#         pass
#     else:
#         tf_adr_eng = shapes[8].text_frame.paragraphs[0]
#         tf_adr_eng.runs[0].text = sheet1.cell(row= r , column= ppt_adr_eng_col).value

#     a = sheet1.cell(row= r , column= ppt_adr_guj_col).value
#     if type(a) == None:
#         pass
#     else:
#         tf_adr_guj = shapes[9].text_frame.paragraphs[0]
#         tf_adr_guj.runs[0].text = sheet1.cell(row= r , column= ppt_adr_guj_col).value
        
#     a = sheet1.cell(row= r , column= ppt_tithi_eng_col).value
#     if type(a) == None:
#         pass
#     else:
#         tf_tithi_eng = shapes[10].text_frame.paragraphs[0]
#         tf_tithi_eng.runs[0].text = sheet1.cell(row= r , column= ppt_tithi_eng_col).value

#     sld_no = sld_no + 1

    

# prs.save(desti_loc_ppt)
# print("Done")
    

